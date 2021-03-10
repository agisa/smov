# -*- coding: utf-8 -*-
"""
Created on Thu Feb 25 08:58:58 2021

@author: elrey
"""

from tkinter import filedialog
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import sys
from openpyxl.utils.cell import column_index_from_string

def get_file_name(name):
    print(name, end = ':\n')
    in_path = filedialog.askopenfilename(title='Chon ' + name)
    print(in_path, end = '\n\n')
    return in_path
#get files path
genname_path = 'AllDataOfGeneratedPowers_18022021.xls' #get_file_name('File AllDataOfGeneratedPowers')
pcsname_path = 'AllDataOfPowerCompanies_18022021.xls' #get_file_name('File AllDataOfPowerCompanies')
subname_path = 'AllDataOfStations_18022021.xls' #get_file_name('File AllDataOfStation')
res_path = 'file_mau.xlsx'
wb = openpyxl.load_workbook(filename=res_path)
ws = wb["cong suat"]

#checking day
subname_day = subname_path.split('.')[0][-8:]
genname_day = genname_path.split('.')[0][-8:]
pcsname_day = pcsname_path.split('.')[0][-8:]
if((subname_day != genname_day) or (subname_day != pcsname_day) or (genname_day != pcsname_day)):
    input('Cac file thong so khong dung hoac khong cung ngay')
    sys.exit()

#defining type of data
data_type = [{'id_col':'NM_ID',     'hourly_para':'Pdc',     'sum_para':'A(dc, gn, dau, td)',   'para_id':'LOAIDL'},
             {'id_col':'DVDK_ID',   'hourly_para':'PHUTAI_P','sum_para':'PHUTAI_A',             'para_id':'LoaiDL'},
             {'id_col':'ID_TBI',    'hourly_para':'P',       'sum_para':'A',                    'para_id':'LOAI_DL'}]
#subname_data = pd.read_excel(subname_path)
#genname_data = pd.read_excel(genname_path)
#pcsname_data = pd.read_excel(pcsname_path)
data = [pd.read_excel(genname_path), pd.read_excel(pcsname_path), pd.read_excel(subname_path)]
process_data = pd.read_excel('list.xlsx')
nrow, ncol = process_data.shape
#hourly data
for i in range(nrow):
    ref_col = data_type[process_data.iloc[i,1]]['id_col']
    ref_para = data_type[process_data.iloc[i,1]]['para_id']
    sum_para = data_type[process_data.iloc[i,1]]['sum_para']
    h_para = data_type[process_data.iloc[i,1]]['hourly_para']
    num_of_ele = process_data.iloc[i, 7]
    dtype = process_data.iloc[i,1]
    row_res = [0]*24
    sum_res = 0
    cur_id_list = str(process_data.iloc[i,8]).split(';')
    res_list = []
    for n in range(num_of_ele):
        #get row with matching id
        row_match = data[dtype][ref_col].astype(str) == cur_id_list[n]
        #get row with matching parameter
        para_match = data[dtype][ref_para] == h_para
        para_sum_match = data[dtype][ref_para] == sum_para
        #adding value
        if dtype == 2:
            tram_match = data[dtype]['MA_TRAM'] == process_data.iloc[i,9]
            matching_row = row_match & para_match & tram_match
            #A tram
            a_tram = data[dtype]['PHAN_LOP'] == 'A_TRAM'
            a_tram_match = a_tram & tram_match
            sum_res = int(round(data[dtype][a_tram_match].loc[:,'H2'])) - int(round(data[dtype][a_tram_match].loc[:,'H1']))
            #matching_sum
        else:
            matching_row = row_match & para_match
            matching_sum = row_match & para_sum_match
            sum_res = data[dtype][matching_sum].loc[:,'H1'].reset_index(drop=True).add(sum_res)
        row_res = data[dtype][matching_row].loc[:,'H1':'H24'].reset_index(drop=True).add(row_res)
        #res_list.append(data[dtype][row_match & para_match].loc[:,'H1':'H24'].reset_index(drop=True))
    #filling hourly data
    row = process_data.iloc[i,2]
    col = process_data.iloc[i,3]
    for r in range(24):
        ws[col + str(row + r)] = int(round(row_res.iloc[:,r]))
    if dtype != 1:
        sum_res *= 1000
    ws[col + str(row + 24)] = int(round(sum_res))
#sum data

    
wb.save('res.xlsx')
os.startfile('res.xlsx', 'open')



#column_index_from_string('A') --> return 1




